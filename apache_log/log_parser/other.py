from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Sum, Count
from .models import LogData


def export_movies_to_xlsx(request):

    ip_queryset = LogData.objects.values("ip").annotate(count=Count('ip')).order_by("-count")[:10]
    unique_ip_queryset = LogData.objects.values('ip').distinct().count()
    methods_queryset = LogData.objects.values("http_method").annotate(count=Count('http_method')).order_by("-count")
    resopnse_size_queryset = LogData.objects.aggregate(Sum('response_size'))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-ip.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Top 10 addresses'

    # Define the titles for columns
    columns = [
        '#',
        'Ip',
        'Total entries'
    ]
    row_num = 1

    # Write titles
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through ip addresses
    cnt = 1
    for item in ip_queryset:
        row_num += 1

        row = [
            cnt,
            item['ip'],
            item['count']
        ]
        cnt += 1

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # write unique_ip
    row_num += 2
    cell = worksheet.cell(row=row_num, column=1)
    cell.value = 'Total unique IP'
    cell = worksheet.cell(row=row_num, column=3)
    cell.value = unique_ip_queryset

    # write methods
    row_num += 1
    cnt = 1
    cell = worksheet.cell(row=row_num, column=1)
    cell.value = 'Http methods'
    for item in methods_queryset:

        row_num += 1

        row = [
            cnt,
            item['http_method'],
            item['count']
        ]
        cnt += 1

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # write total_bytes
    row_num += 2
    cell = worksheet.cell(row=row_num, column=1)
    cell.value = 'Total transmitted bytes'
    cell = worksheet.cell(row=row_num, column=3)
    cell.value = resopnse_size_queryset['response_size__sum']





    workbook.save(response)

    return response

